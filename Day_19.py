from openpyxl import Workbook, load_workbook

wb = Workbook()
sh = wb.active
sh1 = wb.create_sheet('Sheet_2', 2)
sh2 = wb.create_sheet('Sheet_3', 3)

testdata = [['Name', 'Place'], ['Jyoti', 'Pune'], ['Ritesh', 'Bangalore'], ['Hrithik', 'Pune'], ['Aman', 'Delhi']]
testdata1 = [['Name', 'Place','Jyoti', 'Pune']]

for test in testdata:
    sh.append(test)

for test in testdata1:
    sh2.append(test)

for i in range(1, 6):
    for j in range(1, 5):
        sh1.cell(row=i, column=j).value = i +j

wb.save("writing_data_in_excel.xlsx")

# To read data from workbook
wb_to_read = load_workbook("writing_data_in_excel.xlsx")
sheet = wb_to_read['Sheet_2']

print(sheet['A3'].value)
