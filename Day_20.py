# Generate fake data and put it into excel sheet
from openpyxl import Workbook
from faker import Faker

fake_data = Faker()
wb = Workbook()
ws = wb.create_sheet("fake_data", index = 0)

for i in range(1, 21):
    for j in range(1, 4):
        if i == 1 and j == 1:
            ws.cell(1, 1).value = 'Name'
        elif i == 1 and j == 2:
            ws.cell(1, 2).value = 'Email'
        elif i == 1 and j == 3:
            ws.cell(1, 3).value = 'City'
        elif j == 1:
            ws.cell(row=i, column=j).value = fake_data.name()
        elif j == 2:
            ws.cell(row=i, column=j).value = fake_data.email()
        else:
            ws.cell(row=i, column=j).value = fake_data.city()

wb.save("test_data.xlsx")






